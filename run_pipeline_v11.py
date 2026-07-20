#!/usr/bin/env python3
"""
Private Schooling Status Tracker — daily sweep (Google Alerts / Google News edition)
====================================================================================

This is the "sweep." It reads the same searches you have in Google Alerts, pulls the
matching news items, sorts each into an OPENING or a CLOSURE, drops anything already in
your master workbook, and writes the two files the dashboard reads:

    data/candidates.json   (the review-queue cards)
    data/meta.json         (the "Last sweep" timestamp)

The dashboard already knows how to read these on load, every few hours, and when you
click "Run sweep now" — and it ADDS new items to the queue without wiping what's there.

------------------------------------------------------------------------------------
QUICK START
------------------------------------------------------------------------------------
1) Put this file (and, optionally, All_Closures___Openings.xlsx for de-duping) in the
   SAME folder as index.html.
2) (optional, for de-dupe)   pip install openpyxl
3) Run it:                   python3 run_pipeline.py
   -> creates data/candidates.json and data/meta.json
4) Serve the folder (fetch() does NOT work from a file:// path — it must be http):
       python3 -m http.server 8000
   then open  http://localhost:8000/index.html
5) In the dashboard, click "Run sweep now" (or just reload). New candidates appear
   in the review queue and accumulate until you approve/reject them.

------------------------------------------------------------------------------------
RUN IT AUTOMATICALLY EVERY WEEKDAY AT 8:00 AM  (cron)
------------------------------------------------------------------------------------
    crontab -e
    # minute hour * * days-of-week   (1-5 = Mon-Fri).  Use your machine's local time.
    0 8 * * 1-5  cd /full/path/to/the/folder && /usr/bin/python3 run_pipeline.py >> sweep.log 2>&1
(On a server, keep `python3 -m http.server` running, or host the folder on any static
 web host and point cron at it.)

------------------------------------------------------------------------------------
USE YOUR ACTUAL GOOGLE ALERTS FEEDS (exact parity)
------------------------------------------------------------------------------------
By default this queries Google News RSS for the QUERIES below (same idea as your
alerts, nothing to set up). To use your real alerts instead: in Google Alerts, edit an
alert -> "Deliver to" -> "RSS feed", copy the feed URL, and paste it into ALERT_FEEDS.
When ALERT_FEEDS is non-empty, those feeds are used instead of the Google News search.
"""

import json, re, html, urllib.parse, urllib.request
from datetime import datetime, timezone
from xml.etree import ElementTree as ET
from pathlib import Path

# ------------------------------------------------------------------ CONFIG
# Your Google Alerts search terms, exactly as you use them:
QUERIES = [
    "Catholic school closing",
    "COVID school closing",
    "New private school opening",
    "new school opening",
    "Private academy closures",
    "Private school closures",
    "School permanent closing",
]

# OPTIONAL: paste your Google Alerts RSS feed URLs here for exact parity.
# (Google Alerts -> edit alert -> Deliver to: RSS feed -> copy the URL.)
ALERT_FEEDS = [
    # "https://www.google.com/alerts/feeds/00000000000000000000/00000000000000000000",
]

WORKBOOK  = "All_Closures___Openings.xlsx"   # used only for de-duplication (optional)
OUT_DIR   = Path("data")                      # dashboard reads data/candidates.json
DAYS_BACK = 21                                # ignore items older than this many days
UA        = {"User-Agent": "Mozilla/5.0 (school-tracker sweep)"}

# ------------------------------------------------------------------ keyword rules
SCHOOL_HINTS = ("school", "academy", "prep", "montessori", "microschool", "micro-school",
                "micro school", "learning center", "learning academy")
CLOSE_HINTS  = ("clos", "shut", "shutter", "will close", "to close", "permanently")
OPEN_HINTS   = ("open", "opening", "opens", "launch", "new school", "new private",
                "new campus", "new academy", "expand", "expands", "to open", "set to open",
                "will open", "plans to open", "coming to", "under construction",
                "microschool", "micro-school", "micro school")
# almost never what we want (unless clearly a private K-12 school)
STOP_HINTS   = ("university", "board of education", "public school district",
                "school district", "charter school", "community college", "junior college",
                "school board", "county board", "board of trustees", "unified school district",
                "public school", "public schools")
# short abbreviations that only mean something as a whole word (avoid matching
# "isd" inside words like "disdain"):
STOP_WORD_RE = re.compile(r"\bisd\b", re.I)

# U.S.-only scope: skip items that clearly name a non-U.S. place.
# (Heuristic — it removes obvious foreign results; a human still reviews the rest,
#  since a foreign item with no country word in the headline can slip through.)
US_ONLY = True
NON_US_HINTS = (
    "india","delhi","mumbai","bengaluru","bangalore","odisha","kerala","punjab","uttar",
    "uae","dubai","sharjah","abu dhabi","qatar","doha","saudi","riyadh","kuwait","bahrain",
    "united kingdom"," u.k","uk school","london","england","scotland","wales","britain","ireland","dublin",
    "canada","canadian","ontario","toronto","quebec","alberta","manitoba","vancouver","calgary","quinte","napanee",
    "australia","sydney","melbourne","brisbane","new zealand","auckland",
    "nigeria","lagos","abuja","uganda","kenya","nairobi","ghana","accra","south africa","zimbabwe",
    "pakistan","karachi","lahore","bangladesh","dhaka","sri lanka","nepal","afghanistan",
    "singapore","malaysia","kuala lumpur","philippines","manila","indonesia","jakarta","vietnam","thailand",
    "hong kong","china","beijing","shanghai","japan","tokyo","korea","germany","france","spain","italy",
)


def fetch(url):
    with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=45) as r:
        return r.read()


def gnews_url(query):
    return ("https://news.google.com/rss/search?q="
            + urllib.parse.quote(query) + "&hl=en-US&gl=US&ceid=US:en")


def parse_rss(xml_bytes):
    out = []
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return out
    for it in root.iter("item"):
        def g(tag):
            e = it.find(tag)
            return (e.text or "").strip() if e is not None and e.text else ""
        src = it.find("source")
        out.append({
            "title":  html.unescape(g("title")),
            "link":   g("link"),
            "pub":    g("pubDate"),
            "source": (src.text.strip() if src is not None and src.text else ""),
        })
    return out


def parse_date(s):
    for fmt in ("%a, %d %b %Y %H:%M:%S %z", "%a, %d %b %Y %H:%M:%S %Z", "%a, %d %b %Y %H:%M:%S GMT"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def clean_title(title, max_suffix_words=4):
    # Google News formats titles as "Headline - Source" or "Headline | Source";
    # drop the trailing source. Only strip when the trailing segment is SHORT
    # (real source names are 1-4 words: "ITV News", "Austin American-Statesman",
    # "Chronicle Live"). Headlines that just happen to contain " - " as normal
    # punctuation (e.g. "'A terribly sad chapter' - North East private school
    # confirm permanent closure...") have a long trailing segment and are left
    # alone — stripping by position alone (without a length check) silently
    # truncates real headline content.
    best, best_sep_len = -1, 0
    for sep in (" - ", " | "):
        idx = title.rfind(sep)
        if idx > best:
            best, best_sep_len = idx, len(sep)
    if best <= 0:
        return title.strip()
    tail = title[best + best_sep_len:].strip()
    if len(tail.split()) <= max_suffix_words:
        return title[:best].strip()
    return title.strip()


def resolve_gnews_link(url, timeout=20):
    """Resolve a Google News RSS redirect link to the real publisher URL.

    Google News RSS <link> values are opaque redirect pages, not the article
    itself. The real URL is only reachable via an internal batchexecute call
    (there's no plain HTTP redirect or static decode anymore). If Google
    changes this format, this quietly falls back to the original link rather
    than breaking the sweep.
    """
    if "news.google.com" not in url:
        return url
    try:
        page = fetch(url).decode("utf-8", errors="ignore")
        m_id  = re.search(r'data-n-a-id="([^"]+)"', page)
        m_ts  = re.search(r'data-n-a-ts="([^"]+)"', page)
        m_sig = re.search(r'data-n-a-sg="([^"]+)"', page)
        if not (m_id and m_ts and m_sig):
            return url
        article_id, ts, sig = m_id.group(1), int(m_ts.group(1)), m_sig.group(1)
        inner = json.dumps(["garturlreq",
                            [["X", "X", ["X", "X"], None, None, 1, 1, "US:en", None, 1,
                              None, None, None, None, None, 0, 1],
                             "X", "X", 1, [1, 1, 1], 1, 1, None, 0, 0, None, 0],
                            article_id, ts, sig])
        f_req = json.dumps([[["Fbv4je", inner, None, "generic"]]])
        body = urllib.parse.urlencode({"f.req": f_req}).encode()
        req = urllib.request.Request(
            "https://news.google.com/_/DotsSplashUi/data/batchexecute?rpcids=Fbv4je",
            data=body,
            headers={**UA, "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8"},
        )
        with urllib.request.urlopen(req, timeout=timeout) as r:
            resp = r.read().decode("utf-8", errors="ignore")
        # Response is `)]}'` + a JSON array whose 3rd element is itself a
        # JSON-encoded string (with escaped quotes) holding the real URL.
        raw = resp.strip()
        if raw.startswith(")]}'"):
            raw = raw[4:].strip()
        outer = json.loads(raw)
        inner_payload = json.loads(outer[0][2])
        real = inner_payload[1]
        return real if isinstance(real, str) and real.startswith("http") else url
    except Exception as e:
        print(f"  (link resolve failed, keeping Google News link: {e})")
        return url


SCHOOL_NAME_RE = re.compile(
    r"\b((?:[A-Z][\w&.'\u2019-]*|of|the|and|at|St\.?|de)\s+){1,5}?"
    r"(?:High School|Middle School|Elementary School|Preparatory Academy|"
    r"Preparatory School|Learning Academy|Learning Center|Montessori School|"
    r"Micro-?school|Christian School|Catholic School|Academy|School|Prep)"
    r"(?!\s+(?:Board|District|Year|Years|Committee|Bus|Zone|Choice|Boundaries|Boundary))"
    r"\b"
)
NAME_LEADING_FILLER = {"new", "the", "a", "an", "at", "of", "and", "in", "for"}


def extract_school_name(title):
    """Best-effort pull of a specific school name out of a news headline.

    Many headlines ARE just news commentary with no single named school in
    them ("Archdiocese of Santa Fe to close 2 Catholic schools") — in that
    case this returns None and the caller should keep the full headline as
    the candidate name, since inventing a name would be worse than showing
    the headline.
    """
    m = SCHOOL_NAME_RE.search(title)
    if not m:
        return None
    words = m.group(0).strip().split()
    i = 0
    while i < len(words) - 1 and words[i].lower() in NAME_LEADING_FILLER:
        i += 1
    content = words[i:]
    return " ".join(content) if len(content) >= 2 else None


# ------------------------------------------------------------------ article-body name lookup
# RSS headlines frequently name no school ("New School To Open In Lansdale").
# When that happens we open the RESOLVED article and try, in order:
#   1) the outlet's own headline tags (og:title / <h1> / <title>)
#   2) the article BODY, where the subject school is usually the most-repeated
#      named school. A junk/place filter removes nav/social chrome and bare
#      "<City> School" false matches. Best-effort: many sites block bots or
#      bury the name, so returning None (keep the headline) is expected.
SCRIPT_STYLE_RE = re.compile(r'<(script|style)[^>]*>.*?</\1>', re.I | re.S)
TAG_RE = re.compile(r'<[^>]+>')
ARTICLE_TITLE_RES = [
    re.compile(r'<meta[^>]+(?:property|name)=["\'](?:og:title|twitter:title)["\'][^>]+content=["\']([^"\']+)["\']', re.I),
    re.compile(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+(?:property|name)=["\'](?:og:title|twitter:title)["\']', re.I),
    re.compile(r'<h1[^>]*>(.*?)</h1>', re.I | re.S),
    re.compile(r'<title[^>]*>(.*?)</title>', re.I | re.S),
]
ARTICLE_JUNK = ("twitter", "bluesky", "whatsapp", " sms", "email", "facebook",
                "instagram", "tiktok", "top story", "newsletter", "subscribe",
                "sign in", "log in", " menu", "search", "share", "courtesy",
                "showplace", "cookie", "advertisement", "podcast", "copyright",
                "privacy", "terms", "comments", "related", "trending", "most read",
                "sign up")
NAME_LEADING_PLACE = {"new", "top", "the", "a", "this", "that", "downtown",
                      "local", "state", "north", "south", "east", "west", "central"}


def _clean_html_text(s):
    return re.sub(r'\s+', ' ', html.unescape(TAG_RE.sub(' ', s))).strip()


def _plausible_school_name(nm, host):
    """Reject nav/social chrome and bare '<Place> School' matches so the body
    scanner only keeps things that look like a real institution name."""
    low = nm.lower()
    if any(j in low for j in ARTICLE_JUNK):
        return False
    if re.match(r"^[A-Za-z]['’]s\b", nm):        # "A's High School"
        return False
    words = nm.split()
    if len(words) < 2:
        return False
    if words[0].lower() in NAME_LEADING_PLACE:
        return False
    if len(words) == 2:                                # bare "<Place> School/Academy"
        first = words[0].lower().strip(".,")
        host_tokens = set(re.split(r'[.\-]', (host or "").replace("www.", "")))
        if first in STATE_DIVISION or first in host_tokens:
            return False
    return True


def fetch_article_html(url, timeout=15, retries=2):
    """Fetch an article once (first ~600KB) -> (text, host) or (None, None).
    Retries a couple of times so a transient hiccup doesn't blank a card."""
    if not url or not url.startswith("http"):
        return None, None
    host = urllib.parse.urlparse(url).netloc.lower()
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=timeout) as r:
                raw = r.read(600000)
                charset = r.headers.get_content_charset() or "utf-8"
            return raw.decode(charset, "replace"), host
        except Exception:
            if attempt < retries:
                import time; time.sleep(1.5)
    return None, host


def name_from_html(text, host):
    """Pull a specific school name from already-fetched article HTML."""
    if not text:
        return None
    for rx in ARTICLE_TITLE_RES:               # 1) the article's own headline tags
        for m in rx.finditer(text):
            cand = _clean_html_text(m.group(1))
            if cand:
                nm = extract_school_name(clean_title(cand))
                if nm and _plausible_school_name(nm, host):
                    return nm
    body = _clean_html_text(SCRIPT_STYLE_RE.sub(' ', text))   # 2) most-repeated named school
    counts = {}
    for m in SCHOOL_NAME_RE.finditer(body[:8000]):
        nm = extract_school_name(m.group(0))
        if nm and _plausible_school_name(nm, host):
            counts.setdefault(nm, 0)
    if not counts:
        return None
    for nm in counts:
        counts[nm] = len(re.findall(re.escape(nm), body))
    best, _ = sorted(counts.items(), key=lambda kv: (-kv[1], len(kv[0])))[0]
    return best


def fetch_article_name(url, timeout=15):
    text, host = fetch_article_html(url, timeout)
    return name_from_html(text, host)


NON_US_TLDS = (".uk", ".ca", ".au", ".nz", ".ie", ".in", ".za", ".sg", ".my",
               ".ph", ".hk", ".cn", ".jp", ".kr", ".de", ".fr", ".es", ".it",
               ".nl", ".be", ".se", ".no", ".dk", ".fi", ".pl", ".pk", ".bd",
               ".lk", ".np", ".ae", ".sa", ".qa", ".kw", ".bh", ".ng", ".ke",
               ".gh", ".zw", ".id", ".th", ".vn")
NON_US_PATH_SEGMENTS = {"uk", "britain", "england", "scotland", "wales", "northern-ireland"}
NON_US_SOURCE_SUBSTR = (
    "bbc", "itv", "itvx", "the times", "telegraph", "the independent",
    "chronicle live", "dunfermline press", "malverngazette", "christiantoday",
    "irish times", "irish independent", "the scotsman", "belfast telegraph",
    "daily mail", "the sun", "mirror", "sky news", "the guardian",
    "manchester evening news", "birmingham mail", "yorkshire post",
)


def is_non_us(source, resolved_url):
    """Second-pass non-US check using the RESOLVED article URL (domain/path),
    since Google News RSS <source> and headline text alone miss plenty of UK
    stories (e.g. BBC articles live at bbc.com, not bbc.co.uk)."""
    src = (source or "").lower()
    if any(s in src for s in NON_US_SOURCE_SUBSTR):
        return True
    try:
        parsed = urllib.parse.urlparse(resolved_url)
        host = parsed.netloc.lower()
        if any(host.endswith(tld) for tld in NON_US_TLDS):
            return True
        path_segs = {s.lower() for s in parsed.path.split("/") if s}
        if path_segs & NON_US_PATH_SEGMENTS:
            return True
    except Exception:
        pass
    return False


# ------------------------------------------------------------------ state/region (best-effort)
# Conservative: only matches full state names (never 2-letter abbreviations,
# which collide too easily with ordinary words like "OR", "IN", "ME", "HI"),
# and only in the headline + resolved URL PATH — never the source/outlet name,
# since outlet names are unreliable location signals ("Washington Post" isn't
# about Washington state, "Arizona Daily Star" happens to be right but that's
# luck, not a rule).
STATE_DIVISION = {
    "alabama": ("AL", "East South Central"), "alaska": ("AK", "Pacific"),
    "arizona": ("AZ", "Mountain"), "arkansas": ("AR", "West South Central"),
    "california": ("CA", "Pacific"), "colorado": ("CO", "Mountain"),
    "connecticut": ("CT", "New England"), "delaware": ("DE", "South Atlantic"),
    "florida": ("FL", "South Atlantic"), "georgia": ("GA", "South Atlantic"),
    "hawaii": ("HI", "Pacific"), "idaho": ("ID", "Mountain"),
    "illinois": ("IL", "East North Central"), "indiana": ("IN", "East North Central"),
    "iowa": ("IA", "West North Central"), "kansas": ("KS", "West North Central"),
    "kentucky": ("KY", "East South Central"), "louisiana": ("LA", "West South Central"),
    "maine": ("ME", "New England"), "maryland": ("MD", "South Atlantic"),
    "massachusetts": ("MA", "New England"), "michigan": ("MI", "East North Central"),
    "minnesota": ("MN", "West North Central"), "mississippi": ("MS", "East South Central"),
    "missouri": ("MO", "West North Central"), "montana": ("MT", "Mountain"),
    "nebraska": ("NE", "West North Central"), "nevada": ("NV", "Mountain"),
    "new hampshire": ("NH", "New England"), "new jersey": ("NJ", "Middle Atlantic"),
    "new mexico": ("NM", "Mountain"), "new york": ("NY", "Middle Atlantic"),
    "north carolina": ("NC", "South Atlantic"), "north dakota": ("ND", "West North Central"),
    "ohio": ("OH", "East North Central"), "oklahoma": ("OK", "West South Central"),
    "oregon": ("OR", "Pacific"), "pennsylvania": ("PA", "Middle Atlantic"),
    "rhode island": ("RI", "New England"), "south carolina": ("SC", "South Atlantic"),
    "south dakota": ("SD", "West North Central"), "tennessee": ("TN", "East South Central"),
    "texas": ("TX", "West South Central"), "utah": ("UT", "Mountain"),
    "vermont": ("VT", "New England"), "virginia": ("VA", "South Atlantic"),
    "washington": ("WA", "Pacific"), "west virginia": ("WV", "South Atlantic"),
    "wisconsin": ("WI", "East North Central"), "wyoming": ("WY", "Mountain"),
    "district of columbia": ("DC", "South Atlantic"),
}
STATE_RE = re.compile(r"\b(" + "|".join(sorted(STATE_DIVISION, key=len, reverse=True)) + r")\b", re.I)


def url_path_words(url):
    try:
        path = urllib.parse.urlparse(url).path
    except Exception:
        return ""
    segs = [s for s in path.split("/") if s]
    words = []
    for s in segs:
        s = re.sub(r"\.\w{2,5}$", "", s)
        words.append(re.sub(r"[-_]+", " ", s))
    return " ".join(words)


def find_state_region(title, resolved_url, body_text=""):
    """Best-effort (state, region) from the headline + URL path + article body.
    body_text may be raw HTML; we strip it to readable text first (the state
    name usually sits deep past the <head>/scripts). First match wins; returns
    (None, None) rather than guessing when nothing clearly matches."""
    readable = _clean_html_text(SCRIPT_STYLE_RE.sub(' ', body_text)) if body_text else ""
    hay = f"{title} {url_path_words(resolved_url)} {readable[:12000]}".lower()
    m = STATE_RE.search(hay)
    return STATE_DIVISION[m.group(1).lower()] if m else (None, None)


# ------------------------------------------------------------------ NCES address lookup
# NCES Private School Survey is the documented, bot-friendly source for a US
# private school's street address (and it's private-only, so it won't surface
# public schools). It's the 2023-24 survey, so brand-new schools won't appear.
# We fill an address ONLY on a single confident match and tag it "unverified".
STATE_FIPS = {"AL":"01","AK":"02","AZ":"04","AR":"05","CA":"06","CO":"08","CT":"09",
    "DE":"10","DC":"11","FL":"12","GA":"13","HI":"15","ID":"16","IL":"17","IN":"18",
    "IA":"19","KS":"20","KY":"21","LA":"22","ME":"23","MD":"24","MA":"25","MI":"26",
    "MN":"27","MS":"28","MO":"29","MT":"30","NE":"31","NV":"32","NH":"33","NJ":"34",
    "NM":"35","NY":"36","NC":"37","ND":"38","OH":"39","OK":"40","OR":"41","PA":"42",
    "RI":"44","SC":"45","SD":"46","TN":"47","TX":"48","UT":"49","VT":"50","VA":"51",
    "WA":"53","WV":"54","WI":"55","WY":"56"}
ABBR_TO_NAME = {abbr: name.title() for name, (abbr, _div) in STATE_DIVISION.items()}
NCES_GENERIC = {"the","st","st.","saint","school","schools","academy","catholic",
    "christian","high","middle","elementary","preparatory","prep","learning",
    "center","montessori","of","and","college","holy"}
NCES_ROW = re.compile(r'school_detail\.asp\?ID=(\w+)"?>\s*([^<]+?)\s*</a><br /><span>([^<]+)</span>', re.S)


def _nces_core(name):
    words = [w for w in re.sub(r'[^\w\s.]', ' ', name).split()
             if w.lower().strip('.') not in NCES_GENERIC]
    return " ".join(words) if words else re.sub(r'[^\w\s]', ' ', name).strip()


def nces_address(name, state_abbr):
    """Return (address, city) ONLY on a single confident NCES match; None for
    zero results or ambiguous multi-matches. We never guess."""
    fips = STATE_FIPS.get((state_abbr or "").upper())
    core = _nces_core(name or "")
    if not fips or not core:
        return None
    try:
        data = urllib.parse.urlencode({"Search": "1", "SchoolName": core, "State": fips}).encode()
        req = urllib.request.Request(
            "https://nces.ed.gov/surveys/pss/privateschoolsearch/school_list.asp",
            data=data, headers=UA)
        with urllib.request.urlopen(req, timeout=25) as r:
            page = r.read().decode("latin-1", "replace")
    except Exception:
        return None
    rows = []
    for _sid, nm, addr in NCES_ROW.findall(page):
        addr = re.sub(r'\s+', ' ', html.unescape(addr.replace('&nbsp;', ' '))).strip()
        rows.append((html.unescape(nm).strip(), addr))
    if not rows:
        return None
    key = {w.lower() for w in core.split()}
    cand = [r for r in rows if key & set(re.sub(r'[^\w\s]', ' ', r[0]).lower().split())] or rows
    if len(cand) != 1:
        return None                                   # ambiguous -> never guess
    addr = cand[0][1]
    parts = [p.strip() for p in addr.split(",")]
    city = parts[1] if len(parts) >= 2 else ""
    return addr, city


def city_from_body(text, state_abbr):
    """Conservative town pull: only when '<City>, <StateName>' or '<City>, <ST>'
    literally appears in the article. Returns None otherwise (never guesses)."""
    if not text or not state_abbr:
        return None
    body = _clean_html_text(SCRIPT_STYLE_RE.sub(' ', text))[:8000]
    statename = ABBR_TO_NAME.get(state_abbr.upper())
    pats = []
    if statename:
        pats.append(r'([A-Z][a-zA-Z.]+(?:\s[A-Z][a-zA-Z.]+){0,2}),\s+' + re.escape(statename) + r'\b')
    pats.append(r'([A-Z][a-zA-Z.]+(?:\s[A-Z][a-zA-Z.]+){0,2}),\s+' + re.escape(state_abbr.upper()) + r'\b')
    for p in pats:
        m = re.search(p, body)
        if m:
            city = m.group(1).strip()
            if city.lower() not in {"the","a","new","in","of","and","school","academy","this"}:
                return city
    return None


SOCIAL_SOURCES = {"facebook.com", "twitter.com", "x.com", "instagram.com",
                   "tiktok.com", "reddit.com", "youtube.com"}
CALENDAR_HINTS = ("registration open", "start times", "back to school shopping",
                  "first day of school", "school year begins", "school year starts")
STRONG_NEW_HINTS = ("new campus", "new academy", "new private school", "grand opening",
                     "ribbon cutting", "officially opens", "opens its doors",
                     "new school building")


def looks_like_non_event(title, source):
    """Catches items that matched the keyword rules but aren't actually about
    a specific school opening/closing: social-media posts and school-year/
    registration calendar notices that happen to contain "school" + "open"."""
    if (source or "").strip().lower() in SOCIAL_SOURCES:
        return True
    t = title.lower()
    if any(h in t for h in CALENDAR_HINTS) and not any(h in t for h in STRONG_NEW_HINTS):
        return True
    return False


def classify(title):
    t = title.lower()
    if not any(h in t for h in SCHOOL_HINTS):
        return None
    has_stop = any(h in t for h in STOP_HINTS) or bool(STOP_WORD_RE.search(t))
    if has_stop and not any(k in t for k in ("private", "prep", "academy", "catholic")):
        return None
    is_close = any(h in t for h in CLOSE_HINTS)
    is_open  = any(h in t for h in OPEN_HINTS)
    if is_close and not is_open:
        return "closure"
    if is_open and not is_close:
        return "opening"
    if is_close:
        return "closure"   # tie-break: closures are the higher-signal case
    if is_open:
        return "opening"
    return None


def norm(s):
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def load_workbook_names(path):
    names = set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet, col in (("All_closures", 1), ("All_openings", 1)):
            if sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                    if row[col]:
                        names.add(norm(str(row[col])))
    except Exception as e:
        print(f"  (workbook de-dupe skipped: {e})")
    return {n for n in names if len(n) >= 6}   # ignore very short/ambiguous names


# blank record templates matching the dashboard's column schema
OPEN_KEYS  = ["name","link","town","state","region","type","date_reported","opening_date",
              "exp_enrollment","source_links","notes","tuition","tuition_notes","link_b",
              "full_address","tract_mfi","data_year","microschool"]
CLOSE_KEYS = ["name","covid","merging","new_school","town","state","region","type","date_reported",
              "enrollment","enroll_source","year_founded","notes","prefin","link","pct_finaid","link2",
              "tuition","link3","pct_black","enr_black","pct_hispanic","enr_hispanic","pct_asian",
              "enr_asian","pct_white","enr_white","pct_other","enr_other","link4","full_address",
              "tract_mfi","data_year"]


def blank(keys):
    return {k: "" for k in keys}


def main():
    started = datetime.now()
    print("Sweep started:", started.isoformat(timespec="seconds"))

    known = load_workbook_names(WORKBOOK)
    print(f"  workbook: {len(known)} known school names for de-dupe")

    feeds = ALERT_FEEDS if ALERT_FEEDS else [gnews_url(q) for q in QUERIES]
    print(f"  reading {len(feeds)} feed(s): {'your Google Alerts' if ALERT_FEEDS else 'Google News for your queries'}")

    raw = []
    for url in feeds:
        try:
            raw += parse_rss(fetch(url))
        except Exception as e:
            print(f"  feed error: {e}")

    cutoff = datetime.now(timezone.utc).timestamp() - DAYS_BACK * 86400
    seen, candidates, seq = set(), [], 0

    for it in raw:
        title = clean_title(it["title"])
        kind = classify(title)
        if not title or not kind:
            continue
        if US_ONLY:
            hay = (title + " " + it["source"]).lower()
            if any(h in hay for h in NON_US_HINTS):
                print(f"  skip (outside U.S.): {title}")
                continue

        d = parse_date(it["pub"])
        if d:
            ts = (d if d.tzinfo else d.replace(tzinfo=timezone.utc)).timestamp()
            if ts < cutoff:
                continue
            date_reported = d.astimezone().strftime("%Y-%m-%d")
        else:
            date_reported = started.strftime("%Y-%m-%d")

        nn = norm(title)
        if nn in seen:
            continue
        seen.add(nn)

        if any(k in nn for k in known):        # already in the master workbook
            print(f"  skip (already tracked): {title}")
            continue

        extracted_name = extract_school_name(title)
        if not extracted_name and looks_like_non_event(title, it["source"]):
            print(f"  skip (not a specific school event): {title}")
            continue

        real_link = resolve_gnews_link(it["link"])

        # Second-pass non-US check using the resolved article domain/path.
        if US_ONLY and is_non_us(it["source"], real_link):
            print(f"  skip (outside U.S., by source/domain): {title}")
            continue

        # Fetch the article ONCE, then derive name + location from it.
        art_html, art_host = fetch_article_html(real_link)
        if not extracted_name and art_html:
            body_name = name_from_html(art_html, art_host)
            if body_name:
                extracted_name = body_name
                print(f"  name from article: {body_name}")

        needs_name = not extracted_name
        display_name = extracted_name or title

        # Location: state/region from headline + URL + article body (never guessed).
        state, region = find_state_region(title, real_link, art_html or "")
        town, address = "", ""
        if extracted_name and state:               # address via NCES, single confident match only
            hit = nces_address(extracted_name, state)
            if hit:
                address, town = hit
                print(f"  NCES address: {address}")
        # (town is filled only from a verified NCES match above — never guessed
        #  from body text, which produced false positives like photo credits.)

        seq += 1
        if extracted_name:
            note = (f"From the {it['source'] or 'news'} sweep — verify the details "
                     f"before approving. Headline: \"{title}\"")
        else:
            note = (f"From the {it['source'] or 'news'} sweep — headline named no school; "
                     f"add the school name. Headline: \"{title}\"")
        if address:
            note += " Address auto-matched from NCES (2023-24 PSS) — verify."

        if kind == "opening":
            data = blank(OPEN_KEYS)
            data.update(name=display_name, link=real_link, date_reported=date_reported,
                        source_links=it["source"], notes=note,
                        town=town, state=state or "", region=region or "", full_address=address)
        else:
            data = blank(CLOSE_KEYS)
            data.update(name=display_name, link=real_link, date_reported=date_reported,
                        enroll_source=it["source"], notes=note,
                        town=town, state=state or "", region=region or "", full_address=address)

        auto = ["name", "link", "date_reported"]
        for k in ("town", "state", "region", "full_address"):
            if data.get(k):
                auto.append(k)
        candidates.append({"id": f"sw{started:%Y%m%d}{seq:03d}", "kind": kind,
                           "needs_name": needs_name, "headline": title,
                           "auto": auto, "data": data})

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "candidates.json").write_text(json.dumps(candidates, indent=2))

    hour12 = started.strftime("%I").lstrip("0") or "12"
    last_run = started.strftime(f"%a, %b %d %Y &middot; {hour12}:%M %p") + " ET"
    (OUT_DIR / "meta.json").write_text(json.dumps({"last_run": last_run}))

    opens = sum(1 for c in candidates if c["kind"] == "opening")
    closes = len(candidates) - opens
    print(f"  wrote {len(candidates)} candidates ({opens} openings, {closes} closures) -> {OUT_DIR}/candidates.json")
    print("Sweep done.")


if __name__ == "__main__":
    main()
